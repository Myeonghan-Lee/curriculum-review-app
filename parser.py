# -*- coding: utf-8 -*-
"""
parser.py — 교육과정 학점 배당표 엑셀 파서 (v4)

v4 개선사항:
  - 다중 시트 워크북에서 헤더 키워드 점수가 가장 높은 시트를 자동 선택
  - 헤더 탐지 범위를 30행으로 확장하고 정규화 매칭으로 강건성 향상
  - 번호접두사("1)과목"), 줄바꿈("기준\n학점"), 특수문자("교과(군)") 모두 인식
  - 시트별 진단 점수를 오류 메시지에 포함하여 디버깅 용이

출력 인터페이스 (기존 v3와 호환):
  {
    "meta":             {"school": str, "year": int, "version": str},
    "subjects":         [{"구분", "교과군", "과목유형", "과목명",
                          "기준학점", "운영학점",
                          "1-1", "1-2", "2-1", "2-2", "3-1", "3-2",
                          "학기합", "비고", "_row"}],
    "summary":          [{"label", "구분", "운영학점", "이수학점",
                          "필수이수학점", "1-1"...}],
    "notes":            [str],
    "semester_labels":  ["1-1", "1-2", "2-1", "2-2", "3-1", "3-2"],
    "col_map":          {역할: 컬럼번호},
    "diagnostics":      {"selected_sheet", "sheet_scores", "header_range", "data_range"},
  }
"""

from __future__ import annotations
import io
import re
from typing import Dict, List, Tuple, Any
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ---------- 정규화 ----------

def _norm(v: Any) -> str:
    """매칭용 정규형: 공백/줄바꿈/번호접두사/특수문자 제거."""
    if v is None:
        return ""
    s = str(v)
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"^[\(\[]?\d+[\)\]\.]\s*", "", s)
    s = s.replace("·", "").replace("∙", "").replace("・", "")
    return s


HEADER_KEYWORDS = {
    "구분": ["구분"],
    "교과군": ["교과군", "교과(군)", "교과"],
    "과목유형": ["과목유형", "과목구분", "유형"],
    "과목명": ["과목명", "과목"],
    "기준학점": ["기준학점"],
    "운영학점": ["운영학점", "이수단위", "단위"],
    "비고": ["비고"],
    "이수학점": ["이수학점"],
    "필수이수학점": ["필수이수학점", "필수학점"],
    "학년": ["학년"],
    "학기": ["학기"],
}


def _score_header_row(ws: Worksheet, row: int) -> int:
    """특정 행이 헤더일 가능성 점수 (매칭된 고유 키 개수)."""
    matched = set()
    for col in range(1, ws.max_column + 1):
        v = _norm(ws.cell(row, col).value)
        if not v:
            continue
        for key, kws in HEADER_KEYWORDS.items():
            if key in matched:
                continue
            for kw in kws:
                if kw == v or (len(kw) >= 2 and kw in v):
                    matched.add(key)
                    break
    return len(matched)


def _select_best_sheet(wb) -> Tuple[Worksheet, Dict[str, int]]:
    """모든 시트를 스캔하여 헤더 점수가 가장 높은 시트 선택.

    Returns: (선택된 워크시트, {시트명: 최고점수} 진단정보)
    """
    diag: Dict[str, int] = {}
    best_ws, best_score = None, -1
    for sn in wb.sheetnames:
        ws = wb[sn]
        max_row = min(30, ws.max_row)
        sheet_best = 0
        for r in range(1, max_row + 1):
            s = _score_header_row(ws, r)
            if s > sheet_best:
                sheet_best = s
        diag[sn] = sheet_best
        if sheet_best > best_score:
            best_score = sheet_best
            best_ws = ws
    return best_ws, diag


def _find_header_row(ws: Worksheet, scan_rows: int = 30) -> Tuple[int, int]:
    """헤더 시작행과 끝행 자동 탐지. 실패 시 (-1, -1)."""
    max_row = min(scan_rows, ws.max_row)
    best_row, best_score = -1, -1
    for r in range(1, max_row + 1):
        s = _score_header_row(ws, r)
        if s > best_score:
            best_score = s
            best_row = r
    if best_score < 3:
        return -1, -1
    # 다음 1~2행이 학기 행이면 헤더 끝 확장
    end = best_row
    for r in range(best_row + 1, min(best_row + 3, ws.max_row + 1)):
        row_text = " ".join(_norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1))
        if "학기" in row_text and ("1학기" in row_text or "2학기" in row_text):
            end = r
        else:
            break
    return best_row, end


# ---------- 병합 셀 처리 ----------

def _unmerge_and_fill(ws: Worksheet) -> None:
    """병합 셀을 해제하고 좌상단 값을 모든 셀에 복사."""
    for mr in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = mr.bounds
        top_value = ws.cell(min_row, min_col).value
        ws.unmerge_cells(str(mr))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(r, c).value = top_value


# ---------- 메타데이터 ----------

def _extract_metadata(ws: Worksheet, header_start_row: int) -> Dict[str, Any]:
    meta = {"school": None, "year": None, "version": None}
    for r in range(1, header_start_row):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not v:
                continue
            s = str(v)
            if "고등학교" in s and meta["school"] is None:
                m = re.search(r"([가-힣]+(?:여자고등학교|남자고등학교|고등학교))", s)
                if m:
                    meta["school"] = m.group(1)
            m_year = re.search(r"(20\d{2})학년도", s)
            if m_year and meta["year"] is None:
                meta["year"] = int(m_year.group(1))
            if "2022 개정" in s or "2022개정" in s:
                meta["version"] = "2022 개정"
            elif "2015 개정" in s and meta["version"] is None:
                meta["version"] = "2015 개정"
    return meta


# ---------- 컬럼 매핑 ----------

def _build_col_map(ws: Worksheet, hs: int, he: int) -> Tuple[Dict[str, int], List[str]]:
    """헤더 영역을 분석해 {역할: 컬럼번호} 매핑과 학기 라벨 목록 생성."""
    col_map: Dict[str, int] = {}
    used: set = set()
    n_cols = ws.max_column

    # (1) 단일 행 매핑: 긴 키워드 우선
    priority = [
        ("필수이수학점", ["필수이수학점", "필수학점"]),
        ("이수학점", ["이수학점"]),
        ("기준학점", ["기준학점"]),
        ("운영학점", ["운영학점", "이수단위", "단위"]),
        ("과목유형", ["과목유형", "과목구분"]),
        ("과목명", ["과목명", "과목"]),
        ("교과군", ["교과군", "교과(군)", "교과"]),
        ("구분", ["구분"]),
        ("비고", ["비고"]),
    ]
    for key, kws in priority:
        if key in col_map:
            continue
        for r in range(hs, he + 1):
            for c in range(1, n_cols + 1):
                if c in used:
                    continue
                v = _norm(ws.cell(r, c).value)
                if not v:
                    continue
                if key == "과목명" and "유형" in v:
                    continue
                if key == "교과군" and ("유형" in v or "명" in v):
                    continue
                for kw in kws:
                    if kw == v or (len(kw) >= 2 and kw in v):
                        col_map[key] = c
                        used.add(c)
                        break
                if key in col_map:
                    break
            if key in col_map:
                break

    # (2) 학년-학기 매핑
    semester_labels: List[str] = []
    if he > hs:
        for c in range(1, n_cols + 1):
            top = _norm(ws.cell(hs, c).value)
            bot = _norm(ws.cell(he, c).value)
            m_grade = re.search(r"([1-3])학년", top)
            m_sem = re.search(r"([1-2])학기", bot)
            if m_grade and m_sem:
                label = f"{m_grade.group(1)}-{m_sem.group(1)}"
                col_map[label] = c
                if label not in semester_labels:
                    semester_labels.append(label)
    else:
        cur_grade = None
        for c in range(1, n_cols + 1):
            v = _norm(ws.cell(hs, c).value)
            mg = re.search(r"([1-3])학년", v)
            ms = re.search(r"([1-2])학기", v)
            if mg and ms:
                label = f"{mg.group(1)}-{ms.group(1)}"
                col_map[label] = c
                semester_labels.append(label)
            elif mg:
                cur_grade = mg.group(1)
            elif ms and cur_grade:
                label = f"{cur_grade}-{ms.group(1)}"
                col_map[label] = c
                semester_labels.append(label)

    semester_labels.sort()
    return col_map, semester_labels


# ---------- 데이터 추출 ----------

SUMMARY_KEYWORDS = ["합계", "소계", "창의적", "체험활동", "총 이수", "총이수",
                    "유의사항", "학기당", "이수 과목 수", "이수과목수"]


def _is_summary_row(ws: Worksheet, row: int, n_cols: int) -> bool:
    row_text = " ".join(str(ws.cell(row, c).value or "") for c in range(1, n_cols + 1))
    return any(k in row_text for k in SUMMARY_KEYWORDS)


def _to_float(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^\d\.]", "", str(v).strip())
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def _extract_subjects(ws: Worksheet, data_start: int, data_end: int,
                       col_map: Dict[str, int], semester_labels: List[str]) -> List[Dict[str, Any]]:
    subjects: List[Dict[str, Any]] = []
    n_cols = ws.max_column

    for r in range(data_start, data_end + 1):
        if _is_summary_row(ws, r, n_cols):
            continue
        name_col = col_map.get("과목명")
        if not name_col:
            continue
        name_val = ws.cell(r, name_col).value
        if not name_val or not str(name_val).strip():
            continue

        rec: Dict[str, Any] = {
            "_row": r,
            "구분": ws.cell(r, col_map["구분"]).value if "구분" in col_map else None,
            "교과군": ws.cell(r, col_map["교과군"]).value if "교과군" in col_map else None,
            "과목유형": ws.cell(r, col_map["과목유형"]).value if "과목유형" in col_map else None,
            "과목명": str(name_val).strip(),
            "기준학점": _to_float(ws.cell(r, col_map["기준학점"]).value) if "기준학점" in col_map else 0.0,
            "운영학점": _to_float(ws.cell(r, col_map["운영학점"]).value) if "운영학점" in col_map else 0.0,
            "비고": ws.cell(r, col_map["비고"]).value if "비고" in col_map else None,
        }

        total = 0.0
        for label in semester_labels:
            if label in col_map:
                val = _to_float(ws.cell(r, col_map[label]).value)
                rec[label] = val
                total += val
            else:
                rec[label] = 0.0
        rec["학기합"] = total
        if rec["운영학점"] == 0.0 and total > 0:
            rec["운영학점"] = total

        # 비고는 문자열 정리
        if rec["비고"] is not None:
            rec["비고"] = str(rec["비고"]).strip() or None

        subjects.append(rec)
    return subjects


def _extract_summary(ws: Worksheet, data_end: int, col_map: Dict[str, int],
                      semester_labels: List[str]) -> List[Dict[str, Any]]:
    """요약 행(소계/합계/창체/총이수/학기당) 추출."""
    rows: List[Dict[str, Any]] = []
    n_cols = ws.max_column

    for r in range(data_end + 1, ws.max_row + 1):
        text = " ".join(str(ws.cell(r, c).value or "") for c in range(1, n_cols + 1)).strip()
        if not text:
            continue
        if not any(k in text for k in SUMMARY_KEYWORDS):
            continue

        # label 결정
        label_parts = []
        for c in range(1, min(n_cols + 1, 6)):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                sv = str(v).strip()
                if any(k in sv for k in SUMMARY_KEYWORDS) or sv in ("소계", "합계"):
                    label_parts.append(sv)
        label = " ".join(label_parts) if label_parts else text[:30]

        rec: Dict[str, Any] = {
            "_row": r,
            "label": label,
            "구분": ws.cell(r, col_map["구분"]).value if "구분" in col_map else None,
            "교과군": ws.cell(r, col_map["교과군"]).value if "교과군" in col_map else None,
            "운영학점": _to_float(ws.cell(r, col_map["운영학점"]).value) if "운영학점" in col_map else 0.0,
            "이수학점": _to_float(ws.cell(r, col_map["이수학점"]).value) if "이수학점" in col_map else 0.0,
            "필수이수학점": _to_float(ws.cell(r, col_map["필수이수학점"]).value) if "필수이수학점" in col_map else 0.0,
        }
        for label_s in semester_labels:
            rec[label_s] = _to_float(ws.cell(r, col_map[label_s]).value) if label_s in col_map else 0.0
        rows.append(rec)
    return rows


def _extract_notes(ws: Worksheet) -> List[str]:
    """유의사항/비고 텍스트 블록 추출."""
    notes: List[str] = []
    seen = set()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not v:
                continue
            s = str(v).strip()
            if len(s) < 6:
                continue
            if ("유의사항" in s or s.startswith("※") or s.startswith("*") or
                "교차이수" in s or "공동교육과정" in s):
                if s not in seen:
                    seen.add(s)
                    notes.append(s)
    return notes


# ---------- 메인 ----------

def parse_excel(source) -> Dict[str, Any]:
    """엑셀 파일을 파싱해 구조화된 결과 반환.

    Args:
        source: 파일 경로(str) 또는 BytesIO 객체.
    """
    wb = load_workbook(source, data_only=True)

    # 헤더 점수가 가장 높은 시트 자동 선택
    ws, sheet_scores = _select_best_sheet(wb)
    if ws is None:
        raise ValueError("워크북에 시트가 없습니다.")

    _unmerge_and_fill(ws)

    hs, he = _find_header_row(ws)
    if hs < 0:
        diag = ", ".join(f"{k}={v}" for k, v in sheet_scores.items())
        raise ValueError(
            f"헤더 행을 찾지 못했습니다.\n"
            f"  선택 시트: {ws.title}\n"
            f"  시트별 최고 점수: {diag}\n"
            f"  엑셀 상단 30행 안에 '구분/교과/과목/학점/학년' 등 "
            f"헤더 키워드가 3개 이상 포함되어야 합니다."
        )

    meta = _extract_metadata(ws, hs)
    col_map, semester_labels = _build_col_map(ws, hs, he)

    # 데이터 영역 결정
    data_start = he + 1
    data_end = ws.max_row
    for r in range(data_start, ws.max_row + 1):
        if _is_summary_row(ws, r, ws.max_column):
            data_end = r - 1
            break

    subjects = _extract_subjects(ws, data_start, data_end, col_map, semester_labels)
    summary = _extract_summary(ws, data_end, col_map, semester_labels)
    notes = _extract_notes(ws)

    return {
        "meta": meta,
        "subjects": subjects,
        "summary": summary,
        "notes": notes,
        "semester_labels": semester_labels,
        "col_map": col_map,
        "diagnostics": {
            "selected_sheet": ws.title,
            "sheet_scores": sheet_scores,
            "header_range": (hs, he),
            "data_range": (data_start, data_end),
        },
    }
