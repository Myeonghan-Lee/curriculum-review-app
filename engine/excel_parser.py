from openpyxl import load_workbook
import pandas as pd


def fill_merged_cells(ws):

    merged_ranges = list(ws.merged_cells.ranges)

    for merged_range in merged_ranges:

        min_col, min_row, max_col, max_row = merged_range.bounds
        value = ws.cell(min_row, min_col).value
        ws.unmerge_cells(str(merged_range))

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = value



def load_curriculum_excel(file):
    wb = load_workbook(file)
    ws = wb.active
    fill_merged_cells(ws)
    data = ws.values
    df = pd.DataFrame(data)

    return df
